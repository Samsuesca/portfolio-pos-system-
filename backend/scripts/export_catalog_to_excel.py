"""Export DB catalog to Excel files for re-sharing with team.

Genera un Excel por escuela + uno global con el estado actual del catalogo
(post v3_catalog_stab_001 migration). Estos archivos reemplazan los
Guia_Tallas_*.xlsx originales del equipo y se vuelven la fuente canonica
de documentacion (decision #7 del owner).

Estructura del Excel por colegio:
    Sheet "Productos del colegio":
        Garment type | Tallas disponibles | Colores | Precio (rango) | Stock | Tiene imagen

    Sheet "Productos globales" (si aplica):
        idem para globales

Uso:
    python -m backend.scripts.export_catalog_to_excel
    python -m backend.scripts.export_catalog_to_excel --output ./catalog_exports/
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")


def fmt_money(amount) -> str:
    if amount is None:
        return ""
    return f"${int(amount):,}".replace(",", ".")


def fmt_money_range(lo, hi) -> str:
    if lo is None and hi is None:
        return ""
    if lo == hi or hi is None:
        return fmt_money(lo)
    return f"{fmt_money(lo)} - {fmt_money(hi)}"


async def fetch_catalog(session: AsyncSession) -> tuple[list[dict], list[dict], list[dict]]:
    """Returns (schools_list, school_products, global_products) ordered for rendering."""
    schools = (
        await session.execute(
            text(
                """
                SELECT id::text, name, slug
                FROM schools
                WHERE is_active
                ORDER BY name
                """
            )
        )
    ).mappings().all()

    # Productos por escuela
    school_rows = (
        await session.execute(
            text(
                """
                SELECT
                    s.id::text AS school_id,
                    s.name AS school_name,
                    gt.id::text AS gt_id,
                    gt.name AS gt_name,
                    gt.is_active AS gt_active,
                    p.size,
                    p.color,
                    p.gender,
                    p.price,
                    p.is_active AS p_active,
                    COALESCE(i.quantity, 0) AS quantity,
                    COALESCE(imgs.cnt, 0) AS image_count
                FROM garment_types gt
                JOIN schools s ON s.id = gt.school_id
                LEFT JOIN products p ON p.garment_type_id = gt.id
                LEFT JOIN inventory i ON i.product_id = p.id
                LEFT JOIN (
                    SELECT garment_type_id, COUNT(*) AS cnt
                    FROM garment_type_images
                    GROUP BY garment_type_id
                ) imgs ON imgs.garment_type_id = gt.id
                WHERE s.is_active
                ORDER BY s.name, gt.name, p.size NULLS LAST, p.color NULLS LAST
                """
            )
        )
    ).mappings().all()

    # Productos globales
    global_rows = (
        await session.execute(
            text(
                """
                SELECT
                    gt.id::text AS gt_id,
                    gt.name AS gt_name,
                    gt.is_active AS gt_active,
                    p.size,
                    p.color,
                    p.gender,
                    p.price,
                    p.is_active AS p_active,
                    COALESCE(i.quantity, 0) AS quantity,
                    COALESCE(imgs.cnt, 0) AS image_count
                FROM garment_types gt
                LEFT JOIN products p ON p.garment_type_id = gt.id
                LEFT JOIN inventory i ON i.product_id = p.id
                LEFT JOIN (
                    SELECT garment_type_id, COUNT(*) AS cnt
                    FROM garment_type_images
                    GROUP BY garment_type_id
                ) imgs ON imgs.garment_type_id = gt.id
                WHERE gt.school_id IS NULL
                ORDER BY gt.name, p.size NULLS LAST, p.color NULLS LAST
                """
            )
        )
    ).mappings().all()

    return [dict(s) for s in schools], [dict(r) for r in school_rows], [dict(r) for r in global_rows]


def group_by_garment_type(rows: list[dict]) -> list[dict]:
    """Aggregate variant rows by garment_type. One row per gt summarizing variants."""
    groups: dict[str, dict] = {}
    for r in rows:
        gt_id = r["gt_id"]
        if gt_id not in groups:
            groups[gt_id] = {
                "gt_id": gt_id,
                "gt_name": r["gt_name"],
                "gt_active": r["gt_active"],
                "sizes": set(),
                "colors": set(),
                "genders": set(),
                "prices": [],
                "quantities": [],
                "active_variants": 0,
                "total_variants": 0,
                "image_count": r["image_count"],
            }
        g = groups[gt_id]
        if r["size"] is not None:
            g["sizes"].add(r["size"])
            g["total_variants"] += 1
            if r["p_active"]:
                g["active_variants"] += 1
        if r["color"]:
            g["colors"].add(r["color"])
        if r["gender"]:
            g["genders"].add(r["gender"])
        if r["price"] is not None:
            g["prices"].append(r["price"])
        g["quantities"].append(r["quantity"] or 0)

    def sort_sizes(sizes: set) -> list[str]:
        sizes_list = list(sizes)
        def keyfn(s: str):
            try:
                return (0, int(s))
            except ValueError:
                return (1, s)
        return sorted(sizes_list, key=keyfn)

    return [
        {
            **g,
            "sizes": sort_sizes(g["sizes"]),
            "colors": sorted(g["colors"]),
            "genders": sorted(g["genders"]),
            "min_price": min(g["prices"]) if g["prices"] else None,
            "max_price": max(g["prices"]) if g["prices"] else None,
            "total_stock": sum(g["quantities"]),
        }
        for g in groups.values()
    ]


def write_workbook(filepath: Path, title: str, sections: list[tuple[str, list[dict]]]) -> None:
    """Write a workbook with one or more sections.

    Each section is a (label, gt_groups) tuple. Renders a single sheet with sections.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogo"

    row = 1
    # Title
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    ws.cell(row=row, column=1, value=f"Generado: {date.today().isoformat()}").font = Font(italic=True, size=9)
    row += 2

    for section_label, gt_groups in sections:
        # Section header
        ws.cell(row=row, column=1, value=section_label).font = SECTION_FONT
        row += 1

        # Column headers
        headers = ["Producto", "Tallas", "Colores", "Público", "Precio (rango)", "Stock total", "Imagen"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for g in gt_groups:
            ws.cell(row=row, column=1, value=g["gt_name"] + (" (inactivo)" if not g["gt_active"] else ""))
            ws.cell(row=row, column=2, value=", ".join(g["sizes"]) or "(sin variantes)")
            ws.cell(row=row, column=3, value=", ".join(g["colors"]))
            ws.cell(row=row, column=4, value=", ".join(g["genders"]))
            ws.cell(row=row, column=5, value=fmt_money_range(g["min_price"], g["max_price"]))
            ws.cell(row=row, column=6, value=g["total_stock"])
            ws.cell(row=row, column=7, value="Sí" if g["image_count"] > 0 else "—")
            row += 1

        row += 1  # blank line between sections

    # Column widths
    widths = [40, 30, 25, 15, 22, 12, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(filepath)


async def main():
    parser = argparse.ArgumentParser(description="Export catalog to Excel files")
    parser.add_argument("--output", type=Path, default=Path("catalog_exports"))
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)

    async with AsyncSessionLocal() as session:
        schools, school_rows, global_rows = await fetch_catalog(session)

    # Group rows by school
    rows_by_school = defaultdict(list)
    for r in school_rows:
        rows_by_school[r["school_id"]].append(r)

    # Group globals once
    global_groups = group_by_garment_type(global_rows)

    # Generate one Excel per school
    for school in schools:
        sid = school["id"]
        sname = school["name"]
        slug = school["slug"]
        school_groups = group_by_garment_type(rows_by_school.get(sid, []))

        sections = [
            (f"Catálogo de {sname}", school_groups),
            ("Productos globales (compartidos en todos los colegios)", global_groups),
        ]
        safe_slug = slug.replace("/", "_")
        filepath = args.output / f"Catalogo_{safe_slug}.xlsx"
        write_workbook(filepath, f"Catálogo — {sname}", sections)
        print(f"  ✓ {filepath} ({len(school_groups)} escolares + {len(global_groups)} globales)")

    # Generate a global-only Excel
    filepath = args.output / "Catalogo_GLOBALES.xlsx"
    write_workbook(filepath, "Catálogo Global (compartido)", [("Productos globales", global_groups)])
    print(f"  ✓ {filepath} ({len(global_groups)} globales)")

    print(f"\nGenerado {len(schools) + 1} Excels en {args.output}/")


if __name__ == "__main__":
    asyncio.run(main())
