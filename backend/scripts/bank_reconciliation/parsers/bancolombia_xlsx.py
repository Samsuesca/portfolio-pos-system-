"""Parser para extractos Bancolombia formato XLSX.

Estructura observada (verificada en 54089567338_MAR2026.xlsx):
    R2-R4    Información Cliente (header + 1 fila)
    R6-R8    Información General: periodo, tipo cuenta, número, sucursal
    R10-R12  Resumen: saldo anterior, abonos, cargos, saldo actual
    R14      'Movimientos:' (label)
    R15      Header: FECHA | DESCRIPCIÓN | SUCURSAL | DCTO. | VALOR | SALDO
    R16+     Datos. Fecha formato 'D/MM' sin año (inferir del periodo).

Particularidades del valor:
    - String con coma como separador de miles: '34,889,576.33'
    - Prefijo '-' para cargos (salida de plata): '-136.00'
    - '.00' = 0
    - '.81' = 0.81
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterator

import openpyxl


@dataclass
class ParsedHeader:
    period_start: date
    period_end: date
    account_type: str
    account_number: str
    holder_name: str
    opening_balance: Decimal
    closing_balance: Decimal
    total_credits: Decimal
    total_debits: Decimal


@dataclass
class ParsedTransaction:
    transaction_date: date
    raw_description: str
    amount: Decimal           # signed
    running_balance: Decimal | None
    document_number: str | None
    branch: str | None


@dataclass
class BancolombiaStatement:
    header: ParsedHeader
    transactions: list[ParsedTransaction]


def _parse_decimal(value) -> Decimal:
    """'.81' -> 0.81; '-136.00' -> -136; '34,889,576.33' -> 34889576.33."""
    if value is None:
        return Decimal("0")
    s = str(value).strip().replace(",", "")
    if not s or s in (".", "-"):
        return Decimal("0")
    # Casos como '.00' → '0.00'
    if s.startswith("."):
        s = "0" + s
    elif s.startswith("-."):
        s = "-0" + s[1:]
    return Decimal(s)


def _parse_period_date(raw: str) -> date:
    """'2025/12/31' o '2026/03/31' → date."""
    return datetime.strptime(raw, "%Y/%m/%d").date()


def _resolve_year(day: int, month: int, period_start: date, period_end: date) -> int:
    """Asigna año a un mov dado (día, mes) y el rango del periodo.

    Si el rango cruza años, decide según el mes:
      period_start=2025/12/31, period_end=2026/03/31
      → mes 12 → 2025; meses 1-3 → 2026
    """
    if period_start.year == period_end.year:
        return period_start.year
    # Cruza años. Asumimos: meses >= mes_start → año_start; resto → año_end
    if month >= period_start.month:
        return period_start.year
    return period_end.year


def _parse_transaction_date(raw: str, period_start: date, period_end: date) -> date | None:
    """'1/01' → date(year, 1, 1) con año resuelto."""
    if not raw:
        return None
    s = str(raw).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if not m:
        return None
    day = int(m.group(1))
    month = int(m.group(2))
    year = _resolve_year(day, month, period_start, period_end)
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse(path: str | Path) -> BancolombiaStatement:
    """Parsea un extracto Bancolombia XLSX."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no existe: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Table1" not in wb.sheetnames:
        # Por si Bancolombia cambia el nombre: usa la primera hoja
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb["Table1"]

    rows = list(ws.iter_rows(values_only=True))

    # --- HEADER ---
    # R8 (index 7): periodo + tipo + número
    period_row = rows[7]
    period_start = _parse_period_date(str(period_row[0]))
    period_end = _parse_period_date(str(period_row[1]))
    account_type = str(period_row[2]).strip().lower()
    account_number = str(period_row[3]).strip()

    # R4 (index 3): cliente
    holder_name = str(rows[3][0]).strip()

    # R12 (index 11): resumen
    summary_row = rows[11]
    opening_balance = _parse_decimal(summary_row[0])
    total_credits = _parse_decimal(summary_row[1])
    total_debits = _parse_decimal(summary_row[2])
    closing_balance = _parse_decimal(summary_row[3])

    header = ParsedHeader(
        period_start=period_start,
        period_end=period_end,
        account_type=account_type,
        account_number=account_number,
        holder_name=holder_name,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        total_credits=total_credits,
        total_debits=total_debits,
    )

    # --- MOVIMIENTOS ---
    # Header en R15 (index 14), datos desde R16 (index 15)
    transactions: list[ParsedTransaction] = []
    for raw_row in rows[15:]:
        # Defensa: rows vacías al final del sheet
        if not raw_row or all(c is None for c in raw_row):
            continue
        # Defensa: filas que no son movimientos (cualquier label adicional)
        fecha_raw = raw_row[0]
        descripcion = raw_row[1]
        valor = raw_row[4] if len(raw_row) > 4 else None

        # Solo procesamos si parece fila de mov (tiene fecha en formato D/MM)
        txn_date = _parse_transaction_date(fecha_raw, period_start, period_end)
        if txn_date is None or descripcion is None or valor is None:
            continue

        try:
            amount = _parse_decimal(valor)
        except Exception:
            continue

        running_balance = None
        if len(raw_row) > 5 and raw_row[5] is not None:
            try:
                running_balance = _parse_decimal(raw_row[5])
            except Exception:
                running_balance = None

        sucursal = str(raw_row[2]).strip() if len(raw_row) > 2 and raw_row[2] else None
        dcto = str(raw_row[3]).strip() if len(raw_row) > 3 and raw_row[3] else None

        transactions.append(ParsedTransaction(
            transaction_date=txn_date,
            raw_description=str(descripcion).strip(),
            amount=amount,
            running_balance=running_balance,
            document_number=dcto,
            branch=sucursal,
        ))

    return BancolombiaStatement(header=header, transactions=transactions)


def iter_transactions(path: str | Path) -> Iterator[ParsedTransaction]:
    """Conveniencia para streaming."""
    statement = parse(path)
    yield from statement.transactions
