"""Importer idempotente de costos manuscritos (xlsx) al sistema UCR.

Lee los 39 xlsx generados desde cuadernos manuscritos de Consuelo Ríos (workflow
PowerShell+Excel COM operado por el hermano, ver `documentos/Costos/CLAUDE.md`),
los mapea a `cost_component_templates` + `product_cost_components` y los aplica
idempotentemente a la DB.

Decisiones arquitectónicas (ver `docs/v3/formalization/estabilizacion_financiera/
costs-importer-plan-revised.md`):

- **Modelo "un producto por talla"**: `product_cost_components` NO tiene columna
  `size`. Cada talla es un `Product` separado. Componentes variables (fabric,
  collars_cuffs) se asignan amount-por-product matcheando por size; fijos se
  replican a todos los products del garment_type.
- **Mapeo a 8 templates DB canónicos**: fabric, tailoring, embroidery,
  collars_cuffs, labels, bags, thread, other. Conceptos del xlsx (31 únicos)
  agregados a estas categorías vía `CONCEPT_MAP`. Cuando varios conceptos
  mapean al mismo template, se suman.
- **Idempotencia**: claves naturales (product_id, template_id). Update overwrites
  el amount existente, no inserta duplicado.
- **Fórmulas xlsx calculadas manualmente**: no depende de `data_only=True` (que
  requiere que Excel haya abierto el archivo post-edit). Calcula
  `costo_total_corte = (cm/100) * precio_metro` y `costo_unit = total / unidades`.
- **Tallas Pendiente**: skipean, anotan en gaps report. No crean component con
  amount=0 (engañoso).
- **Bloques mixed sizes** (Chompa Pumarejo: numéricas 6-16 + letras S-XXL): cada
  bloque tiene sus propios componentes (cierre numérico vs letras, confección
  numérica vs letras). Aplican solo a los products cuya size matchea el bloque.
- **Jomber cross-school**: archivo Jomber_Pumarejo_Pinal_Caracas.xlsx aplica a 3
  schools. Procesa la primera copia, ignora duplicadas.

Uso:
    cd backend
    venv/bin/python -m scripts.import_costs_from_xlsx                  # dry-run con zip default
    venv/bin/python -m scripts.import_costs_from_xlsx --source <path>  # dry-run con path
    venv/bin/python -m scripts.import_costs_from_xlsx --commit         # persiste
    venv/bin/python -m scripts.import_costs_from_xlsx --commit --report /tmp/report.md
"""
from __future__ import annotations

import argparse
import asyncio
import logging
import re
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import AsyncSessionLocal
from app.models.product import (
    CostComponentTemplate,
    GarmentType,
    Product,
    ProductCostComponent,
)
from app.models.school import School
from app.utils.timezone import get_colombia_now_naive

logging.basicConfig(level=logging.INFO, format="%(message)s")
logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)
logger = logging.getLogger("import_costs")

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ZIP = REPO_ROOT / "documentos" / "Costos" / "COSTOS-20260516T232636Z-3-001.zip"


# ---------------------------------------------------------------------------
# Mapeos hardcoded (school code / garment / concept)
# ---------------------------------------------------------------------------

# Códigos en filenames → SQL LIKE pattern del school.name.
# JIGL aplica a 2 colegios (Jardín Gota De Leche + Jardin Infantil Fe y Alegria).
SCHOOL_MAP: dict[str, list[str]] = {
    "Comfama":      ["Comfama"],
    "BuenComienzo": ["Buen Comienzo"],
    "Caracas":      ["Caracas"],
    "FHB":          ["Felix Henao"],
    "HAG":          ["Héctor Abad"],
    "JDLCP":        ["Juan De La Cruz"],
    "JIGL":         ["Jardin Infantil Fe y Alegria", "Jardín Gota De Leche"],
    "MJC":          ["Manuel José Caycedo"],
    "Pinal":        ["El Pinal"],
    "Pumarejo":     ["Alfonso López Pumarejo"],
    # Jomber: archivo cross-school, se trata aparte en main()
    "Jomber":       ["Alfonso López Pumarejo", "Caracas", "El Pinal"],
}

# Prenda del filename → patterns ILIKE de garment_type.name (case-insensitive).
# Múltiples patterns matchean varias variantes (Comfama "Camiseta Amarillo/Azul/...").
# Algunos colegios usan "Camisa" en vez de "Camiseta" (HAG, JIGL, Jardín Gota), por eso %camis%.
GARMENT_MAP: dict[str, list[str]] = {
    "Camiseta":          ["%camis%"],
    "Camiseta_Algodon":  ["%algod%"],
    "Camiseta_Diario":   ["%camiseta%diario%", "CAMISETA"],  # FHB/MJC/etc lo llaman solo CAMISETA
    "Camiseta_Fisica":   ["%fisica%", "%física%", "Camisa Fisica", "CAMISETA FISICA"],
    "Sudadera":          ["%sudadera%"],
    "Chompa":            ["%chompa%"],
    "Chompa_Azul":       ["%chompa%azul%"],
    "Chompa Gris":       ["%chompa%gris%"],  # nota: con espacio en filename
    "Delantal":          ["%delantal%"],
    "Pumarejo_Pinal_Caracas": ["%jumper%"],  # archivo Jomber_* = Jumper en DB
    "Pinal_Caracas":          ["%jumper%"],  # variantes del split del filename Jomber_*
    "Jomber":                 ["%jumper%"],
}

# Patterns negativos: excluir matches cuyo nombre contenga estos substrings.
# Ej.: prenda "Camiseta" no debe matchear "Camisa de física" (que tiene su propio xlsx).
# Para "Chompa" NO se filtran Azul/Gris porque solo Caracas tiene variantes específicas
# y NO tiene xlsx "Chompa" simple — no hay colisión.
GARMENT_NEGATIVE_MAP: dict[str, list[str]] = {
    "Camiseta":          ["fisica", "física", "diario", "algod"],
}

# Concepto normalizado (lower, sin tildes excepto "ñ"/"ó") → código de template DB.
# Múltiples conceptos al mismo template se SUMAN (e.g., "marquilla logo"+"talla c/u" → labels).
CONCEPT_MAP: dict[str, str] = {
    # → labels (marquillas / etiquetas)
    "marquilla logo":            "labels",
    "marquilla logo y talla":    "labels",
    "marquilla logo+talla":      "labels",
    "talla c/u":                 "labels",

    # → tailoring (confección + corte)
    "corte":                     "tailoring",
    "confeccion":                "tailoring",
    "confeccion (incluye corte)": "tailoring",

    # → bags
    "bolsa + cinta":             "bags",
    "bolsa cinta":               "bags",
    "bolsa y cinta":             "bags",

    # → embroidery
    "bordado":                   "embroidery",
    "escudo bordado":            "embroidery",

    # → collars_cuffs (cuellos/puños/perillas/entretelas)
    "cuello":                    "collars_cuffs",
    "cuello + puno":             "collars_cuffs",
    "cuello letra tejida plana": "collars_cuffs",
    "perilla":                   "collars_cuffs",
    "entretela perilla":         "collars_cuffs",
    "entretela perilla (40cm)":  "collars_cuffs",

    # → other (insumos diversos: cierres, broches, botones, resorte)
    "broche":                    "other",
    "broches":                   "other",
    "5 broches":                 "other",
    "botones (5 x $150)":        "other",
    "boton":                     "other",
    "cierre":                    "other",
    "cierre separable":          "other",
    "cierres (3 x $450)":        "other",
    "resorte":                   "other",
}

# Conceptos por bloque (Chompa Pumarejo: cierres distintos para tallas numéricas vs letras).
# Aplican solo a products cuya size matchee el bloque.
CONCEPT_MAP_NUMERIC_BLOCK: dict[str, str] = {
    "cierre separable numericas":   "other",
    "confeccion tallas numericas":  "tailoring",
}
CONCEPT_MAP_LETTERS_BLOCK: dict[str, str] = {
    "cierre separable letras":     "other",
    "confeccion tallas letras":    "tailoring",
}

# Tipo de tela en Hoja "Telas" → categoría:
#   "principal" / "secundaria" → tela principal del fabric
#   "complemento" / "rib" / "sesgo" / "cuello" / "forro" / "colores" → collars_cuffs
TELA_PRINCIPAL_TYPES = {"principal", "secundaria"}
TELA_COMPLEMENTO_TYPES = {"complemento", "rib", "sesgo", "cuello", "forro", "colores"}


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class TelaRow:
    """Fila de la hoja Telas."""
    name: str
    tipo: str           # principal/secundaria/complemento/etc
    colores: str | None
    precio_por_metro: Decimal
    row_idx: int        # 4-indexed (D4, D5, D6 referenciados por fórmulas Consumo Tela)


@dataclass
class InsumoRow:
    """Fila de la hoja Insumos (concepto + valor)."""
    concept_raw: str
    concept_norm: str
    value: Decimal
    notes: str | None
    template_code: str | None  # mapeado vía CONCEPT_MAP
    applies_to_block: str | None = None  # "numeric" | "letters" | None (todos)


@dataclass
class SizeFabricCost:
    """Costo de tela calculado por talla."""
    size: str                       # "2", "10", "S", "XL"
    cm: Decimal | None              # cm de tela por corte (None si Pendiente)
    unidades_por_corte: Decimal | None
    tela_ref_idx: int               # qué tela de la hoja Telas (4=principal, 5=secundaria...)
    costo_unitario: Decimal | None  # calculado: (cm/100)*precio/unidades. None si Pendiente
    is_pendiente: bool = False
    block: str | None = None        # "numeric" | "letters" | None


@dataclass
class ComplementoCost:
    """Costo de tela complemento (RIB/sesgo/forro). Flat o variable."""
    is_flat: bool
    flat_amount: Decimal | None = None
    per_size: list[SizeFabricCost] = field(default_factory=list)


@dataclass
class XlsxData:
    """Parsed structure de un xlsx de costos."""
    filename: str
    school_code: str
    prenda_code: str
    telas: list[TelaRow]
    insumos: list[InsumoRow]
    sizes: list[SizeFabricCost]
    complemento: ComplementoCost | None
    gaps: list[str]


@dataclass
class ApplyResult:
    file: str
    schools_matched: list[str]
    garment_types_matched: list[tuple[str, str]]  # (school_name, gt_name)
    products_touched: int
    components_inserted: int
    components_updated: int
    components_skipped: int
    errors: list[str]
    gaps: list[str]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(s: object) -> str:
    """Lowercase, strip, collapse spaces. Mantiene tildes para matching consistente."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def to_decimal(v: object) -> Decimal | None:
    """Convierte a Decimal preservando enteros y skipping fórmulas/strings."""
    if v is None or isinstance(v, str):
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def classify_size(size: str) -> str:
    """Devuelve 'numeric' si la size es int (2, 6, 10, 16), 'letters' si S/M/L/XL/XXL."""
    s = size.strip().upper()
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return "numeric"
    if s in {"S", "M", "L", "XL", "XXL", "XXXL"}:
        return "letters"
    return "unknown"


def map_concept(concept_norm: str) -> tuple[str | None, str | None]:
    """Devuelve (template_code, block). Block in {numeric, letters, None}."""
    if concept_norm in CONCEPT_MAP_NUMERIC_BLOCK:
        return CONCEPT_MAP_NUMERIC_BLOCK[concept_norm], "numeric"
    if concept_norm in CONCEPT_MAP_LETTERS_BLOCK:
        return CONCEPT_MAP_LETTERS_BLOCK[concept_norm], "letters"
    if concept_norm in CONCEPT_MAP:
        return CONCEPT_MAP[concept_norm], None
    return None, None


def parse_filename(fname: str) -> tuple[str | None, str | None]:
    """Extrae (school_code, prenda_code) del filename. None,None si no matchea."""
    # Jomber_Pumarejo_Pinal_Caracas.xlsx → school="Jomber", prenda="Pumarejo_Pinal_Caracas"
    if fname.startswith("Jomber_") or fname.startswith(" Jomber_"):
        return "Jomber", fname.replace(" Jomber_", "").replace("Jomber_", "").replace(".xlsx", "")
    # Colegio_<School>_<Prenda>.xlsx
    m = re.match(r"Colegio_([^_]+)_(.+)\.xlsx$", fname)
    if m:
        return m.group(1), m.group(2)
    return None, None


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _parse_telas(ws: Worksheet) -> list[TelaRow]:
    telas: list[TelaRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        name, tipo, colores, precio = (row + (None,) * 4)[:4]
        if not name:
            continue
        precio_dec = to_decimal(precio)
        if precio_dec is None:
            continue
        telas.append(TelaRow(
            name=str(name).strip(),
            tipo=normalize(tipo),
            colores=str(colores).strip() if colores else None,
            precio_por_metro=precio_dec,
            row_idx=idx,
        ))
    return telas


def _parse_insumos(ws: Worksheet) -> list[InsumoRow]:
    insumos: list[InsumoRow] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        concept, value, notes = (row + (None,) * 3)[:3]
        if not concept:
            continue
        c_norm = normalize(concept)
        if c_norm.startswith("total"):
            break  # stop al llegar a TOTAL INSUMOS
        value_dec = to_decimal(value)
        if value_dec is None:
            continue
        template, block = map_concept(c_norm)
        insumos.append(InsumoRow(
            concept_raw=str(concept).strip(),
            concept_norm=c_norm,
            value=value_dec,
            notes=str(notes).strip() if notes else None,
            template_code=template,
            applies_to_block=block,
        ))
    return insumos


def _calc_fabric_cost(cm: Decimal, unidades: Decimal, telas: list[TelaRow], ref_idx: int) -> Decimal | None:
    """Calcula (cm/100) * precio_metro / unidades."""
    tela = next((t for t in telas if t.row_idx == ref_idx), None)
    if not tela or unidades <= 0:
        return None
    total_corte = (cm / Decimal(100)) * tela.precio_por_metro
    return (total_corte / unidades).quantize(Decimal("0.01"))


def _parse_size_row(
    row: tuple, telas: list[TelaRow], default_tela_idx: int, block: str | None,
) -> SizeFabricCost | None:
    """Parsea una fila de Consumo Tela. Devuelve None si la fila no es de datos."""
    a, b, c, d, e = (row + (None,) * 5)[:5]
    if a is None:
        return None
    size = str(a).strip()
    if not size or any(kw in size.lower() for kw in ("tallas numericas", "tallas letras", "complemento",
                                                       "sesgo", "rib", "forro", "tela complemento")):
        return None  # header de bloque, no fila de datos

    pendiente = any(isinstance(v, str) and "pendiente" in v.lower() for v in (b, c, d, e))
    if pendiente:
        return SizeFabricCost(
            size=size, cm=None, unidades_por_corte=None,
            tela_ref_idx=default_tela_idx, costo_unitario=None,
            is_pendiente=True, block=block,
        )

    cm = to_decimal(b)
    unidades = to_decimal(c)
    if cm is None or unidades is None:
        return None

    # detectar tela_ref_idx desde fórmula en D (e.g., "=(B4/100)*Telas!$D$4" → ref=4)
    tela_ref = default_tela_idx
    if isinstance(d, str):
        m = re.search(r"\$D\$(\d+)", d)
        if m:
            tela_ref = int(m.group(1))

    # Si E (costo c/u) es un número directo (no fórmula), confiar en el valor del xlsx
    # — la dueña puede haber ajustado manualmente con precio efectivo distinto al nominal.
    e_direct = to_decimal(e)
    if e_direct is not None and e_direct > 0:
        costo = e_direct.quantize(Decimal("0.01"))
    else:
        costo = _calc_fabric_cost(cm, unidades, telas, tela_ref)
    return SizeFabricCost(
        size=size, cm=cm, unidades_por_corte=unidades,
        tela_ref_idx=tela_ref, costo_unitario=costo,
        is_pendiente=False, block=block,
    )


def _parse_consumo_tela(ws: Worksheet, telas: list[TelaRow]) -> tuple[list[SizeFabricCost], ComplementoCost | None]:
    """
    Parsea Consumo Tela. Devuelve (sizes_principales, complemento).

    El complemento puede ser:
      - Flat: una fila después de "TELA COMPLEMENTO (RIB)" con size="Todas" y costo único.
      - Variable: tabla por talla bajo "SESGO" o "TELA COMPLEMENTO".
    """
    sizes: list[SizeFabricCost] = []
    complemento_per_size: list[SizeFabricCost] = []
    complemento_flat: Decimal | None = None

    in_complemento = False
    in_block_numeric = False
    in_block_letters = False
    complemento_tela_idx = 5  # default: segunda fila de Telas (D5)

    for row in ws.iter_rows(min_row=4, values_only=True):
        a = row[0]
        if a is None and all(v is None for v in row):
            continue
        a_str = str(a).strip().lower() if a is not None else ""

        # Detectar headers de sección
        if "tallas numericas" in a_str:
            in_block_numeric = True
            in_block_letters = False
            in_complemento = False
            continue
        if "tallas letras" in a_str:
            in_block_numeric = False
            in_block_letters = True
            in_complemento = False
            continue
        if any(kw in a_str for kw in ("tela complemento", "sesgo", "rib", "forro")):
            in_complemento = True
            in_block_numeric = False
            in_block_letters = False
            # detectar idx de tela complemento desde alguna referencia en la fila
            for v in row:
                if isinstance(v, str):
                    m = re.search(r"\$D\$(\d+)", v)
                    if m:
                        complemento_tela_idx = int(m.group(1))
                        break
            continue
        # Header secundario en complemento (e.g., "Talla | Sesgo total (cm) | ...")
        if in_complemento and a_str == "talla":
            continue

        block = None
        if in_block_numeric:
            block = "numeric"
        elif in_block_letters:
            block = "letters"

        default_tela_idx = telas[0].row_idx if telas else 4
        if in_complemento:
            default_tela_idx = complemento_tela_idx

        size_cost = _parse_size_row(row, telas, default_tela_idx, block)
        if size_cost is None:
            continue

        if in_complemento:
            # Caso flat: size="Todas"
            if size_cost.size.lower() == "todas":
                if size_cost.costo_unitario is not None:
                    complemento_flat = size_cost.costo_unitario
            else:
                complemento_per_size.append(size_cost)
        else:
            sizes.append(size_cost)

    if complemento_flat is not None:
        complemento = ComplementoCost(is_flat=True, flat_amount=complemento_flat)
    elif complemento_per_size:
        complemento = ComplementoCost(is_flat=False, per_size=complemento_per_size)
    else:
        complemento = None

    return sizes, complemento


def parse_xlsx(path: Path) -> XlsxData:
    """Carga y parsea un xlsx de costos. Lanza ValueError si la estructura es inválida."""
    wb = load_workbook(path, data_only=False)
    school_code, prenda_code = parse_filename(path.name)
    if school_code is None:
        raise ValueError(f"filename no reconocido: {path.name}")

    gaps: list[str] = []

    if "Telas" not in wb.sheetnames:
        raise ValueError("falta hoja Telas")
    telas = _parse_telas(wb["Telas"])
    if not telas:
        raise ValueError("hoja Telas vacía")

    if "Insumos" not in wb.sheetnames:
        raise ValueError("falta hoja Insumos")
    insumos = _parse_insumos(wb["Insumos"])

    # Detectar conceptos no mapeados (gap report)
    for ins in insumos:
        if ins.template_code is None:
            gaps.append(f"concepto no mapeado: '{ins.concept_raw}' ({ins.value})")

    if "Consumo Tela" not in wb.sheetnames:
        raise ValueError("falta hoja Consumo Tela")
    sizes, complemento = _parse_consumo_tela(wb["Consumo Tela"], telas)

    pendientes = [s.size for s in sizes if s.is_pendiente]
    if pendientes:
        gaps.append(f"tallas pendientes: {', '.join(pendientes)}")

    return XlsxData(
        filename=path.name,
        school_code=school_code,
        prenda_code=prenda_code or "",
        telas=telas,
        insumos=insumos,
        sizes=sizes,
        complemento=complemento,
        gaps=gaps,
    )


# ---------------------------------------------------------------------------
# DB matching + apply
# ---------------------------------------------------------------------------

async def _find_schools(db: AsyncSession, school_code: str) -> list[School]:
    patterns = SCHOOL_MAP.get(school_code, [])
    if not patterns:
        return []
    matched: list[School] = []
    seen_ids: set = set()
    for pattern in patterns:
        stmt = select(School).where(School.name.ilike(f"%{pattern}%"))
        result = await db.execute(stmt)
        for sch in result.scalars().all():
            if sch.id not in seen_ids:
                matched.append(sch)
                seen_ids.add(sch.id)
    return matched


async def _find_garment_types(
    db: AsyncSession, schools: list[School], prenda_code: str,
) -> list[GarmentType]:
    if not schools:
        return []
    patterns = GARMENT_MAP.get(prenda_code, [])
    if not patterns:
        # fallback: prenda_code mismo como pattern
        patterns = [f"%{prenda_code.lower()}%"]
    negative_keywords = [k.lower() for k in GARMENT_NEGATIVE_MAP.get(prenda_code, [])]

    school_ids = [s.id for s in schools]
    matched: list[GarmentType] = []
    seen_ids: set = set()
    for pattern in patterns:
        stmt = select(GarmentType).where(
            GarmentType.school_id.in_(school_ids),
            GarmentType.name.ilike(pattern),
        )
        result = await db.execute(stmt)
        for gt in result.scalars().all():
            if gt.id in seen_ids:
                continue
            gt_lower = gt.name.lower()
            if any(neg in gt_lower for neg in negative_keywords):
                continue
            matched.append(gt)
            seen_ids.add(gt.id)
    return matched


async def _find_or_create_template(
    db: AsyncSession, garment_type_id, code: str, is_variable: bool,
) -> CostComponentTemplate:
    """Busca template por (garment_type_id, code). Crea si no existe."""
    stmt = select(CostComponentTemplate).where(
        CostComponentTemplate.garment_type_id == garment_type_id,
        CostComponentTemplate.code == code,
    )
    result = await db.execute(stmt)
    tpl = result.scalar_one_or_none()
    if tpl is not None:
        return tpl

    # crear minimal — el resto de templates ya existen por migración
    tpl = CostComponentTemplate(
        garment_type_id=garment_type_id,
        code=code,
        name=_template_display_name(code),
        is_variable=is_variable,
        display_order=999,
        is_active=True,
    )
    db.add(tpl)
    await db.flush()
    return tpl


def _template_display_name(code: str) -> str:
    return {
        "fabric": "Tela",
        "tailoring": "Confección",
        "embroidery": "Bordado",
        "collars_cuffs": "Cuellos/Puños",
        "labels": "Marquillas",
        "bags": "Bolsas",
        "thread": "Hilos",
        "other": "Otros",
    }.get(code, code.title())


async def _upsert_component(
    db: AsyncSession, product_id, template_id, amount: Decimal, notes: str | None,
) -> str:
    """Upsert por (product_id, template_id). Devuelve 'inserted' | 'updated' | 'skipped'."""
    stmt = select(ProductCostComponent).where(
        ProductCostComponent.product_id == product_id,
        ProductCostComponent.template_id == template_id,
    )
    result = await db.execute(stmt)
    existing = result.scalar_one_or_none()
    now = get_colombia_now_naive()

    if existing is None:
        db.add(ProductCostComponent(
            product_id=product_id,
            template_id=template_id,
            amount=amount,
            notes=notes,
            created_at=now,
            updated_at=now,
        ))
        return "inserted"
    if existing.amount == amount and (existing.notes or None) == (notes or None):
        return "skipped"
    existing.amount = amount
    existing.notes = notes
    existing.updated_at = now
    return "updated"


def _aggregate_insumos(insumos: list[InsumoRow]) -> dict[tuple[str, str | None], tuple[Decimal, str]]:
    """
    Suma insumos por (template_code, block). Devuelve {(code, block): (amount, notes)}.
    Block None = aplica a todos los products. Block 'numeric'/'letters' = solo a esa categoría.
    """
    agg: dict[tuple[str, str | None], tuple[Decimal, list[str]]] = {}
    for ins in insumos:
        if ins.template_code is None:
            continue
        key = (ins.template_code, ins.applies_to_block)
        cur_amount, cur_notes = agg.get(key, (Decimal("0"), []))
        cur_amount += ins.value
        cur_notes.append(f"{ins.concept_raw}={ins.value}")
        agg[key] = (cur_amount, cur_notes)
    return {k: (amt, "; ".join(notes)) for k, (amt, notes) in agg.items()}


def _insumos_for_block(
    agg: dict[tuple[str, str | None], tuple[Decimal, str]],
    block: str,
) -> dict[str, tuple[Decimal, str]]:
    """Colapsa agg a {template_code: (amount, notes)} sumando block=None + block=match."""
    result: dict[str, tuple[Decimal, list[str]]] = {}
    for (code, ins_block), (amt, notes) in agg.items():
        if ins_block is not None and ins_block != block:
            continue
        cur_amt, cur_notes = result.get(code, (Decimal("0"), []))
        cur_amt += amt
        cur_notes.append(notes)
        result[code] = (cur_amt, cur_notes)
    return {code: (amt, "; ".join(notes)) for code, (amt, notes) in result.items()}


async def apply_xlsx_to_db(db: AsyncSession, data: XlsxData) -> ApplyResult:
    """Aplica los componentes parseados a todos los products matched."""
    result = ApplyResult(
        file=data.filename,
        schools_matched=[],
        garment_types_matched=[],
        products_touched=0,
        components_inserted=0,
        components_updated=0,
        components_skipped=0,
        errors=[],
        gaps=list(data.gaps),
    )

    schools = await _find_schools(db, data.school_code)
    if not schools:
        result.errors.append(f"school code '{data.school_code}' no match en DB")
        return result
    result.schools_matched = [s.name for s in schools]

    garment_types = await _find_garment_types(db, schools, data.prenda_code)
    if not garment_types:
        result.errors.append(
            f"garment '{data.prenda_code}' no match para schools {result.schools_matched}"
        )
        return result
    school_by_id = {s.id: s for s in schools}
    result.garment_types_matched = [
        (school_by_id[gt.school_id].name, gt.name) for gt in garment_types
    ]

    insumos_agg = _aggregate_insumos(data.insumos)

    for gt in garment_types:
        # cargar products del garment_type
        prod_stmt = select(Product).where(
            Product.garment_type_id == gt.id, Product.is_active.is_(True),
        )
        prod_result = await db.execute(prod_stmt)
        products = prod_result.scalars().all()
        if not products:
            continue

        # templates para este garment_type
        tpl_fabric = await _find_or_create_template(db, gt.id, "fabric", is_variable=True)
        tpl_complement = await _find_or_create_template(db, gt.id, "collars_cuffs", is_variable=False)

        insumo_templates: dict[str, CostComponentTemplate] = {}
        for (code, _block), _ in insumos_agg.items():
            if code not in insumo_templates:
                insumo_templates[code] = await _find_or_create_template(
                    db, gt.id, code, is_variable=False,
                )

        for product in products:
            psize = (product.size or "").strip()
            psize_class = classify_size(psize)
            touched = False
            # Set para garantizar máximo 1 upsert por (product, template) en esta corrida
            seen_templates_this_product: set = set()

            # --- 1. Fabric (variable por talla) ---
            fabric_cost = next(
                (s for s in data.sizes if s.size == psize and not s.is_pendiente
                 and s.costo_unitario is not None),
                None,
            )
            if fabric_cost is not None:
                op = await _upsert_component(
                    db, product.id, tpl_fabric.id, fabric_cost.costo_unitario,
                    notes=f"xlsx={data.filename}; cm={fabric_cost.cm}; unid={fabric_cost.unidades_por_corte}",
                )
                _bump(result, op)
                seen_templates_this_product.add(tpl_fabric.id)
                touched = True

            # --- 2. Complemento (tela) ---
            if data.complemento is not None and tpl_complement.id not in seen_templates_this_product:
                comp_amount: Decimal | None = None
                if data.complemento.is_flat:
                    comp_amount = data.complemento.flat_amount
                else:
                    match = next(
                        (s for s in data.complemento.per_size
                         if s.size == psize and not s.is_pendiente
                         and s.costo_unitario is not None),
                        None,
                    )
                    if match:
                        comp_amount = match.costo_unitario
                if comp_amount is not None:
                    op = await _upsert_component(
                        db, product.id, tpl_complement.id, comp_amount,
                        notes=f"xlsx={data.filename}; tela complemento",
                    )
                    _bump(result, op)
                    seen_templates_this_product.add(tpl_complement.id)
                    touched = True

            # --- 3. Insumos: colapsar block-None + block-matched a un único set ---
            block_for_product = psize_class if psize_class in ("numeric", "letters") else "numeric"
            effective_insumos = _insumos_for_block(insumos_agg, block_for_product)
            for code, (amount, notes) in effective_insumos.items():
                tpl = insumo_templates[code]
                if tpl.id in seen_templates_this_product:
                    continue
                op = await _upsert_component(
                    db, product.id, tpl.id, amount,
                    notes=f"xlsx={data.filename}; {notes}",
                )
                _bump(result, op)
                seen_templates_this_product.add(tpl.id)
                touched = True

            # Flush por product para que el siguiente product vea los inserts previos
            await db.flush()

            if touched:
                result.products_touched += 1

    return result


def _bump(result: ApplyResult, op: str) -> None:
    if op == "inserted":
        result.components_inserted += 1
    elif op == "updated":
        result.components_updated += 1
    elif op == "skipped":
        result.components_skipped += 1


# ---------------------------------------------------------------------------
# Reporter
# ---------------------------------------------------------------------------

def render_report(results: list[ApplyResult], mode: str, source: str) -> str:
    lines = [
        f"# Costs Import Report — {mode}",
        f"",
        f"- Source: `{source}`",
        f"- Files processed: {len(results)}",
        f"- Total components inserted: {sum(r.components_inserted for r in results)}",
        f"- Total components updated:  {sum(r.components_updated for r in results)}",
        f"- Total components skipped:  {sum(r.components_skipped for r in results)}",
        f"- Total products touched:    {sum(r.products_touched for r in results)}",
        f"- Files with errors:         {sum(1 for r in results if r.errors)}",
        f"- Files with gaps:           {sum(1 for r in results if r.gaps)}",
        f"",
        f"## Detalle por archivo",
        f"",
    ]
    for r in sorted(results, key=lambda x: x.file):
        status = "✓" if not r.errors else "✗"
        lines.append(f"### {status} {r.file}")
        if r.errors:
            for e in r.errors:
                lines.append(f"  - ERROR: {e}")
        else:
            lines.append(f"  - Schools: {', '.join(r.schools_matched)}")
            lines.append(f"  - Garment types matched: {len(r.garment_types_matched)}")
            for s, g in r.garment_types_matched:
                lines.append(f"      • {s} / {g}")
            lines.append(
                f"  - Products touched: {r.products_touched} | "
                f"inserted: {r.components_inserted} | updated: {r.components_updated} | "
                f"skipped: {r.components_skipped}"
            )
        if r.gaps:
            lines.append(f"  - Gaps:")
            for g in r.gaps:
                lines.append(f"      • {g}")
        lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def resolve_source(source: str | None) -> tuple[Path, bool]:
    """Devuelve (dir_con_xlsx, is_temp). Extrae zip si es necesario."""
    if source:
        src_path = Path(source)
        if src_path.is_dir():
            return src_path, False
        if src_path.is_file() and src_path.suffix == ".zip":
            tmp = Path(tempfile.mkdtemp(prefix="ucr-costs-"))
            with zipfile.ZipFile(src_path) as zf:
                zf.extractall(tmp)
            return tmp, True
        raise ValueError(f"--source debe ser un dir o un .zip: {src_path}")
    # default: extrae el zip canónico
    if not DEFAULT_ZIP.exists():
        raise FileNotFoundError(f"default zip no encontrado: {DEFAULT_ZIP}")
    tmp = Path(tempfile.mkdtemp(prefix="ucr-costs-"))
    with zipfile.ZipFile(DEFAULT_ZIP) as zf:
        zf.extractall(tmp)
    return tmp, True


def collect_xlsx(root: Path) -> list[Path]:
    """Lista los xlsx evitando duplicados de Jomber_*.xlsx (que aparecen 3 veces)."""
    seen_basenames: set[str] = set()
    files: list[Path] = []
    for p in sorted(root.rglob("*.xlsx")):
        base = p.name.strip()
        # los 3 Jomber_Pumarejo_Pinal_Caracas.xlsx son idénticos; procesar una sola vez
        if base.startswith("Jomber_") and base in seen_basenames:
            continue
        if base.startswith(" Jomber_") and base.strip() in seen_basenames:
            continue
        seen_basenames.add(base)
        files.append(p)
    return files


async def main(args: argparse.Namespace) -> int:
    mode = "COMMIT (persiste)" if args.commit else "DRY-RUN (rollback)"
    logger.info("=== Import Costs From XLSX — %s ===", mode)

    src_dir, is_temp = resolve_source(args.source)
    logger.info("Source dir: %s", src_dir)

    xlsx_files = collect_xlsx(src_dir)
    logger.info("Archivos a procesar: %d", len(xlsx_files))

    results: list[ApplyResult] = []

    async with AsyncSessionLocal() as db:
        for path in xlsx_files:
            try:
                data = parse_xlsx(path)
            except Exception as e:
                results.append(ApplyResult(
                    file=path.name,
                    schools_matched=[], garment_types_matched=[],
                    products_touched=0, components_inserted=0,
                    components_updated=0, components_skipped=0,
                    errors=[f"parse error: {e}"], gaps=[],
                ))
                logger.warning("✗ %s: parse error: %s", path.name, e)
                continue

            try:
                r = await apply_xlsx_to_db(db, data)
            except Exception as e:
                await db.rollback()
                results.append(ApplyResult(
                    file=path.name,
                    schools_matched=[], garment_types_matched=[],
                    products_touched=0, components_inserted=0,
                    components_updated=0, components_skipped=0,
                    errors=[f"apply error: {type(e).__name__}: {e}"], gaps=data.gaps,
                ))
                logger.warning("✗ %s: apply error: %s", path.name, e)
                continue

            results.append(r)
            status = "✓" if not r.errors else "✗"
            logger.info(
                "%s %s — schools=%d gts=%d prods=%d ins=%d upd=%d skip=%d",
                status, path.name, len(r.schools_matched), len(r.garment_types_matched),
                r.products_touched, r.components_inserted, r.components_updated,
                r.components_skipped,
            )

        if args.commit:
            await db.commit()
            logger.info("=== COMMIT — cambios persistidos ===")
        else:
            await db.rollback()
            logger.info("=== DRY-RUN — rollback, re-ejecuta con --commit ===")

    report = render_report(results, mode, str(src_dir))
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(report, encoding="utf-8")
    logger.info("Report: %s", report_path)

    total_inserted = sum(r.components_inserted for r in results)
    total_updated = sum(r.components_updated for r in results)
    total_skipped = sum(r.components_skipped for r in results)
    total_errors = sum(1 for r in results if r.errors)
    logger.info(
        "Totales: inserted=%d updated=%d skipped=%d errors=%d",
        total_inserted, total_updated, total_skipped, total_errors,
    )

    if is_temp:
        logger.info("(dir temporal: %s — borrar manualmente si querés)", src_dir)

    return 0 if total_errors == 0 else 1


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--source", default=None,
        help="Path a dir con xlsx, o a un .zip. Default: documentos/Costos/COSTOS-*.zip"
    )
    parser.add_argument("--commit", action="store_true", help="Persiste cambios.")
    parser.add_argument(
        "--report", default="/tmp/ucr-costos-report.md",
        help="Path del markdown de report.",
    )
    parsed = parser.parse_args()
    sys.exit(asyncio.run(main(parsed)))
